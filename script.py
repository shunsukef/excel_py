import openpyxl
import os

output_file = open('output.txt','w')

wb = openpyxl.load_workbook('sample.xlsx')
#print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet1')
#print(sheet.title)

#c = sheet['A1']
#print(c.value)

for i in range (1,4):
    wire_a = sheet.cell(row = i ,column =1).value
    wire_b = sheet.cell(row = i ,column =2).value
    output_file.write(str(wire_a) + ' = ' + str(wire_b) + '\n')
    
output_file.close()